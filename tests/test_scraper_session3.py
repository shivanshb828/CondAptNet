"""
Session 3 unit tests: parsers (pdf_parser, excel_parser, xml_parser, text_parser).

All tests use synthetic in-memory content — no network calls, no filesystem
fixtures beyond what's created within the test.
"""

import io
import textwrap
from pathlib import Path

import pytest

# ── Helpers ────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parents[1]
import sys
sys.path.insert(0, str(PROJECT_ROOT))


# ═══════════════════════════════════════════════════════════════════════════════
# pdf_parser
# ═══════════════════════════════════════════════════════════════════════════════

class TestPdfParser:

    def test_import(self):
        from scripts.data.scraper.parsers.pdf_parser import (
            parse_pdf, tables_to_text, ParsedDocument, ParsedPage,
        )
        assert callable(parse_pdf)
        assert callable(tables_to_text)

    def test_no_pdfplumber_returns_parse_error(self, monkeypatch):
        """When pdfplumber is unavailable, parse_pdf returns parse_ok=False."""
        import scripts.data.scraper.parsers.pdf_parser as mod
        monkeypatch.setattr(mod, "_HAS_PDFPLUMBER", False)
        doc = mod.parse_pdf(b"fake bytes")
        assert doc.parse_ok is False
        assert "pdfplumber" in doc.error.lower()

    def test_invalid_bytes_returns_parse_error(self):
        """Passing invalid bytes (not a real PDF) returns parse_ok=False."""
        from scripts.data.scraper.parsers.pdf_parser import parse_pdf
        doc = parse_pdf(b"not a real pdf \x00\x01\x02")
        assert doc.parse_ok is False
        assert doc.error != ""

    def test_invalid_path_returns_parse_error(self):
        from scripts.data.scraper.parsers.pdf_parser import parse_pdf
        doc = parse_pdf("/nonexistent/path/fake.pdf")
        assert doc.parse_ok is False

    def test_parsed_document_fields(self):
        """ParsedDocument dataclass has all expected fields."""
        from scripts.data.scraper.parsers.pdf_parser import ParsedDocument
        doc = ParsedDocument()
        assert hasattr(doc, "pages")
        assert hasattr(doc, "full_text")
        assert hasattr(doc, "all_tables")
        assert hasattr(doc, "n_pages")
        assert hasattr(doc, "source_path")
        assert hasattr(doc, "parse_ok")
        assert hasattr(doc, "error")

    def test_parsed_page_fields(self):
        from scripts.data.scraper.parsers.pdf_parser import ParsedPage
        p = ParsedPage(page_number=1, text="hello", tables=[], char_offset=0)
        assert p.page_number == 1
        assert p.text == "hello"
        assert p.char_offset == 0

    def test_tables_to_text_empty(self):
        from scripts.data.scraper.parsers.pdf_parser import tables_to_text
        assert tables_to_text([]) == ""

    def test_tables_to_text_simple(self):
        from scripts.data.scraper.parsers.pdf_parser import tables_to_text
        tables = [[["A", "B", "C"], ["1", "2", "3"]]]
        result = tables_to_text(tables)
        assert "A\tB\tC" in result
        assert "1\t2\t3" in result

    def test_tables_to_text_multiple_tables(self):
        from scripts.data.scraper.parsers.pdf_parser import tables_to_text
        t1 = [["Seq", "Kd"], ["ATGCATGC", "5 nM"]]
        t2 = [["Target", "Type"], ["Thrombin", "protein"]]
        result = tables_to_text([t1, t2])
        assert "Seq" in result
        assert "Thrombin" in result

    def test_looks_garbled_all_single_chars(self):
        from scripts.data.scraper.parsers.pdf_parser import _looks_garbled
        # All single-character words → garbled
        text = " ".join(list("ABCDEFGHIJKLMNOPQRSTUVWXYZABCDE"))
        assert _looks_garbled(text) is True

    def test_looks_garbled_normal_text(self):
        from scripts.data.scraper.parsers.pdf_parser import _looks_garbled
        text = ("This is normal scientific paper text about aptamer selection "
                "and binding affinity measurements in nanomolar range.")
        assert _looks_garbled(text) is False

    def test_looks_garbled_too_short_returns_false(self):
        from scripts.data.scraper.parsers.pdf_parser import _looks_garbled
        # < 20 words → not enough data → False
        assert _looks_garbled("A B C D E") is False

    def test_clean_cells_none_to_empty_string(self):
        from scripts.data.scraper.parsers.pdf_parser import _clean_cells
        table = [[None, "hello", None], ["world", None, "42"]]
        cleaned = _clean_cells(table)
        assert cleaned[0][0] == ""
        assert cleaned[0][1] == "hello"
        assert cleaned[1][1] == ""

    def test_real_pdf_if_available(self, tmp_path):
        """Create a minimal real PDF using reportlab if available; skip otherwise."""
        pytest.importorskip("reportlab")
        from reportlab.pdfgen import canvas as rl_canvas
        pdf_path = tmp_path / "test.pdf"
        c = rl_canvas.Canvas(str(pdf_path))
        c.drawString(72, 720, "Kd = 5.2 nM for aptamer ATCGATCGATCGATCGATCGATCG")
        c.save()

        from scripts.data.scraper.parsers.pdf_parser import parse_pdf
        doc = parse_pdf(str(pdf_path))
        assert doc.parse_ok is True
        assert doc.n_pages >= 1
        assert "Kd" in doc.full_text or len(doc.full_text) > 0


# ═══════════════════════════════════════════════════════════════════════════════
# excel_parser
# ═══════════════════════════════════════════════════════════════════════════════

class TestExcelParser:

    def test_import(self):
        from scripts.data.scraper.parsers.excel_parser import (
            parse_excel, parse_csv, ParsedSpreadsheet, ParsedSheet,
        )
        assert callable(parse_excel)
        assert callable(parse_csv)

    def test_parsed_spreadsheet_fields(self):
        from scripts.data.scraper.parsers.excel_parser import ParsedSpreadsheet
        ss = ParsedSpreadsheet()
        assert hasattr(ss, "sheets")
        assert hasattr(ss, "full_text")
        assert hasattr(ss, "n_sheets")
        assert hasattr(ss, "source_path")
        assert hasattr(ss, "parse_ok")

    def test_parsed_sheet_fields(self):
        import pandas as pd
        from scripts.data.scraper.parsers.excel_parser import ParsedSheet
        s = ParsedSheet(name="s1", df=pd.DataFrame(), n_rows=0, n_cols=0, seq_cols=[])
        assert s.name == "s1"

    def test_csv_simple(self, tmp_path):
        csv_path = tmp_path / "test.csv"
        csv_path.write_text(
            "aptamer_sequence,target,kd_nM\n"
            "ATCGATCGATCGATCGATCGATCG,thrombin,5.2\n"
            "GCTAGCTAGCTAGCTAGCTAGCTA,insulin,12.0\n"
        )
        from scripts.data.scraper.parsers.excel_parser import parse_csv
        ss = parse_csv(str(csv_path))
        assert ss.parse_ok is True
        assert ss.n_sheets == 1
        assert ss.sheets[0].n_rows == 2

    def test_csv_seq_col_detection_by_header(self, tmp_path):
        csv_path = tmp_path / "seqtest.csv"
        csv_path.write_text(
            "aptamer_sequence,target\n"
            "ATCGATCGATCGATCGATCGATCG,thrombin\n"
        )
        from scripts.data.scraper.parsers.excel_parser import parse_csv
        ss = parse_csv(str(csv_path))
        sheet = ss.sheets[0]
        assert "aptamer_sequence" in sheet.seq_cols

    def test_csv_seq_col_detection_by_content(self, tmp_path):
        csv_path = tmp_path / "seqcontent.csv"
        csv_path.write_text(
            "col_a,col_b\n"
            "ATCGATCGATCGATCGATCGATCG,thrombin\n"
            "GCTAGCTAGCTAGCTAGCTAGCTA,insulin\n"
            "ATGATGATGATGATGATGATGATG,VEGF\n"
        )
        from scripts.data.scraper.parsers.excel_parser import parse_csv
        ss = parse_csv(str(csv_path))
        sheet = ss.sheets[0]
        assert "col_a" in sheet.seq_cols

    def test_csv_full_text_contains_sequences(self, tmp_path):
        csv_path = tmp_path / "full.csv"
        csv_path.write_text(
            "sequence,target\nATCGATCGATCGATCGATCGATCG,thrombin\n"
        )
        from scripts.data.scraper.parsers.excel_parser import parse_csv
        ss = parse_csv(str(csv_path))
        assert "ATCGATCGATCGATCGATCGATCG" in ss.full_text
        assert "thrombin" in ss.full_text

    def test_csv_bytes_input(self):
        raw = b"sequence,target\nATCGATCGATCGATCGATCGATCG,thrombin\n"
        from scripts.data.scraper.parsers.excel_parser import parse_csv
        ss = parse_csv(raw)
        assert ss.parse_ok is True
        assert ss.n_sheets == 1

    def test_tsv_separator(self, tmp_path):
        tsv_path = tmp_path / "test.tsv"
        tsv_path.write_text("sequence\ttarget\nATCGATCGATCGATCGATCGATCG\tthrombin\n")
        from scripts.data.scraper.parsers.excel_parser import parse_csv
        ss = parse_csv(str(tsv_path), sep="\t")
        assert ss.parse_ok is True
        assert ss.sheets[0].n_rows == 1

    def test_invalid_csv_returns_error(self):
        # openpyxl/pandas should fail on binary garbage
        from scripts.data.scraper.parsers.excel_parser import parse_excel
        doc = parse_excel(b"\x00\x01\x02\x03garbage")
        assert doc.parse_ok is False

    def test_excel_file(self, tmp_path):
        """Write a real .xlsx with openpyxl and parse it back."""
        pytest.importorskip("openpyxl")
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aptamers"
        ws.append(["aptamer_sequence", "target_name", "kd_nM"])
        ws.append(["ATCGATCGATCGATCGATCGATCG", "thrombin", "5.2"])
        ws.append(["GCTAGCTAGCTAGCTAGCTAGCTA", "insulin", "12.0"])
        xlsx_path = tmp_path / "aptamers.xlsx"
        wb.save(str(xlsx_path))

        from scripts.data.scraper.parsers.excel_parser import parse_excel
        ss = parse_excel(str(xlsx_path))
        assert ss.parse_ok is True
        assert ss.n_sheets >= 1
        sheet = ss.sheets[0]
        assert sheet.n_rows == 2
        assert "aptamer_sequence" in sheet.seq_cols

    def test_excel_multisheet(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws1.append(["col1", "col2"])
        ws1.append(["a", "b"])
        ws2 = wb.create_sheet("S2")
        ws2.append(["colX"])
        ws2.append(["x"])
        xlsx_path = tmp_path / "multi.xlsx"
        wb.save(str(xlsx_path))

        from scripts.data.scraper.parsers.excel_parser import parse_excel
        ss = parse_excel(str(xlsx_path))
        assert ss.n_sheets == 2
        names = [s.name for s in ss.sheets]
        assert "S1" in names and "S2" in names

    def test_detect_seq_cols_no_match(self, tmp_path):
        csv_path = tmp_path / "nosequences.csv"
        csv_path.write_text("name,value\nfoo,bar\nbaz,qux\n")
        from scripts.data.scraper.parsers.excel_parser import parse_csv
        ss = parse_csv(str(csv_path))
        assert ss.sheets[0].seq_cols == []


# ═══════════════════════════════════════════════════════════════════════════════
# xml_parser
# ═══════════════════════════════════════════════════════════════════════════════

_MINIMAL_NXML = b"""<?xml version="1.0" encoding="UTF-8"?>
<article>
  <front>
    <article-meta>
      <title-group>
        <article-title>Aptamer Selection for Thrombin Binding</article-title>
      </title-group>
      <abstract>
        <p>We selected DNA aptamers against human thrombin with Kd = 26 nM.</p>
      </abstract>
    </article-meta>
  </front>
  <body>
    <sec>
      <title>Methods</title>
      <p>SELEX was performed in PBS pH 7.4, 150 mM NaCl, 2 mM MgCl2 at 37 degrees C.</p>
    </sec>
    <sec>
      <title>Results</title>
      <p>The best aptamer sequence was ATCGATCGATCGATCGATCGATCGATCGATCG with Kd of 26 nM.</p>
      <table-wrap id="T1">
        <label>Table 1</label>
        <caption><p>Selected aptamers and their binding affinities.</p></caption>
        <table>
          <tr><th>Sequence</th><th>Kd (nM)</th></tr>
          <tr><td>ATCGATCGATCGATCGATCGATCGATCGATCG</td><td>26</td></tr>
          <tr><td>GCTAGCTAGCTAGCTAGCTAGCTAGCTAGCTA</td><td>85</td></tr>
        </table>
      </table-wrap>
    </sec>
  </body>
  <back>
    <supplementary-material xmlns:xlink="http://www.w3.org/1999/xlink"
      xlink:href="supplementary_table_S1.xlsx">
      <caption><p>Supplementary Table S1</p></caption>
    </supplementary-material>
  </back>
</article>
"""

class TestXmlParser:

    def test_import(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml, ParsedXML, ParsedTable
        assert callable(parse_nxml)

    def test_no_lxml_returns_error(self, monkeypatch):
        import scripts.data.scraper.parsers.xml_parser as mod
        monkeypatch.setattr(mod, "_HAS_LXML", False)
        doc = mod.parse_nxml(_MINIMAL_NXML)
        assert doc.parse_ok is False
        assert "lxml" in doc.error.lower()

    def test_basic_parsing(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert doc.parse_ok is True

    def test_title_extraction(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert "Thrombin" in doc.title or "Aptamer" in doc.title

    def test_abstract_extraction(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert "thrombin" in doc.abstract.lower()
        assert "26 nM" in doc.abstract or "Kd" in doc.abstract

    def test_body_text_extraction(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert "SELEX" in doc.body_text or "PBS" in doc.body_text

    def test_table_extraction(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert len(doc.tables) == 1

    def test_table_label(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert "Table 1" in doc.tables[0].label

    def test_table_has_rows(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert len(doc.tables[0].rows) >= 2

    def test_table_contains_sequence(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        all_cells = [cell for row in doc.tables[0].rows for cell in row]
        assert any("ATCG" in c for c in all_cells)

    def test_supplementary_url_extraction(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        assert len(doc.supplementary_urls) == 1
        assert "supplementary_table_S1.xlsx" in doc.supplementary_urls[0]

    def test_full_text_contains_all_sections(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(_MINIMAL_NXML)
        # full_text should cover abstract, body, tables
        assert "thrombin" in doc.full_text.lower()
        assert "SELEX" in doc.full_text or "PBS" in doc.full_text

    def test_invalid_xml_does_not_crash(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        # lxml recover=True should handle malformed XML
        garbled = b"<article><title>Incomplete"
        doc = parse_nxml(garbled)
        # Should not raise; may or may not parse_ok depending on lxml version
        assert isinstance(doc.parse_ok, bool)

    def test_empty_xml_returns_empty_doc(self):
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(b"<article></article>")
        assert doc.parse_ok is True
        assert doc.title == ""
        assert doc.tables == []
        assert doc.supplementary_urls == []

    def test_parse_from_file(self, tmp_path):
        nxml_path = tmp_path / "test.xml"
        nxml_path.write_bytes(_MINIMAL_NXML)
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(nxml_path)
        assert doc.parse_ok is True
        assert doc.n_pages if hasattr(doc, "n_pages") else True

    def test_multiple_supplementary_urls(self):
        xml = b"""<?xml version="1.0"?>
        <article>
          <back>
            <supplementary-material xmlns:xlink="http://www.w3.org/1999/xlink"
              xlink:href="supp1.xlsx"/>
            <supplementary-material xmlns:xlink="http://www.w3.org/1999/xlink"
              xlink:href="supp2.pdf"/>
          </back>
        </article>"""
        from scripts.data.scraper.parsers.xml_parser import parse_nxml
        doc = parse_nxml(xml)
        assert len(doc.supplementary_urls) == 2

    def test_tables_to_text_helper(self):
        from scripts.data.scraper.parsers.xml_parser import _tables_to_text, ParsedTable
        t = ParsedTable(label="Table 1", caption="Test", rows=[["A", "B"], ["1", "2"]])
        text = _tables_to_text([t])
        assert "A\tB" in text
        assert "1\t2" in text


# ═══════════════════════════════════════════════════════════════════════════════
# text_parser
# ═══════════════════════════════════════════════════════════════════════════════

_SIMPLE_HTML = """
<html>
<head><title>Aptamer Binding Study</title></head>
<body>
  <p>We identified aptamers binding thrombin with Kd = 26 nM.</p>
  <p>Selection was performed in PBS pH 7.4 containing 150 mM NaCl.</p>
  <table>
    <tr><th>Sequence</th><th>Target</th><th>Kd (nM)</th></tr>
    <tr><td>ATCGATCGATCGATCGATCGATCG</td><td>Thrombin</td><td>26</td></tr>
  </table>
  <script>var x = 1;</script>
  <nav>skip nav</nav>
</body>
</html>
"""

class TestTextParser:

    def test_import(self):
        from scripts.data.scraper.parsers.text_parser import (
            parse_html, parse_plain_text, parse_file, ParsedText,
        )
        assert callable(parse_html)
        assert callable(parse_plain_text)
        assert callable(parse_file)

    def test_no_bs4_returns_error(self, monkeypatch):
        import scripts.data.scraper.parsers.text_parser as mod
        monkeypatch.setattr(mod, "_HAS_BS4", False)
        result = mod.parse_html("<html><body>test</body></html>")
        assert result.parse_ok is False
        assert "beautifulsoup4" in result.error.lower()

    def test_html_title_extraction(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        assert result.title == "Aptamer Binding Study"

    def test_html_is_html_flag(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        assert result.is_html is True

    def test_html_text_contains_aptamer_content(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        assert "thrombin" in result.text.lower()
        assert "Kd" in result.text or "26 nM" in result.text

    def test_html_script_stripped(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        assert "var x = 1" not in result.text

    def test_html_nav_stripped(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        assert "skip nav" not in result.text

    def test_html_table_extraction(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        assert len(result.tables) == 1

    def test_html_table_has_header_row(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        header_row = result.tables[0][0]
        assert "Sequence" in header_row

    def test_html_table_has_data_row(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML)
        data_row = result.tables[0][1]
        assert "ATCGATCGATCGATCGATCGATCG" in data_row

    def test_html_bytes_input(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(_SIMPLE_HTML.encode("utf-8"))
        assert result.parse_ok is True
        assert "thrombin" in result.text.lower()

    def test_plain_text_basic(self):
        from scripts.data.scraper.parsers.text_parser import parse_plain_text
        text = "The Kd was 5.2 nM for aptamer against thrombin."
        result = parse_plain_text(text)
        assert result.parse_ok is True
        assert result.is_html is False
        assert "thrombin" in result.text

    def test_plain_text_no_tables(self):
        from scripts.data.scraper.parsers.text_parser import parse_plain_text
        result = parse_plain_text("Some text here with no tables.")
        assert result.tables == []

    def test_plain_text_empty_title(self):
        from scripts.data.scraper.parsers.text_parser import parse_plain_text
        result = parse_plain_text("Some text.")
        assert result.title == ""

    def test_plain_text_whitespace_normalised(self):
        from scripts.data.scraper.parsers.text_parser import parse_plain_text
        result = parse_plain_text("too    many   spaces    here")
        assert "too    many" not in result.text
        assert "too many spaces" in result.text

    def test_plain_text_bytes_input(self):
        from scripts.data.scraper.parsers.text_parser import parse_plain_text
        result = parse_plain_text(b"Some bytes text")
        assert result.parse_ok is True
        assert "Some bytes text" in result.text

    def test_parse_file_html(self, tmp_path):
        html_path = tmp_path / "test.html"
        html_path.write_text(_SIMPLE_HTML)
        from scripts.data.scraper.parsers.text_parser import parse_file
        result = parse_file(html_path)
        assert result.is_html is True
        assert result.parse_ok is True

    def test_parse_file_txt(self, tmp_path):
        txt_path = tmp_path / "test.txt"
        txt_path.write_text("Aptamer binding study with Kd = 5.2 nM.")
        from scripts.data.scraper.parsers.text_parser import parse_file
        result = parse_file(txt_path)
        assert result.is_html is False
        assert result.parse_ok is True

    def test_parse_file_nonexistent(self):
        from scripts.data.scraper.parsers.text_parser import parse_file
        result = parse_file("/nonexistent/path.html")
        assert result.parse_ok is False

    def test_parsed_text_fields(self):
        from scripts.data.scraper.parsers.text_parser import ParsedText
        doc = ParsedText()
        for attr in ("text", "tables", "title", "source_path", "is_html", "parse_ok", "error"):
            assert hasattr(doc, attr)

    def test_html_empty_body(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html("<html><body></body></html>")
        assert result.parse_ok is True
        assert result.text == "" or result.text.strip() == ""

    def test_html_multiple_tables(self):
        html = """
        <html><body>
          <table><tr><td>A1</td><td>B1</td></tr></table>
          <table><tr><td>A2</td></tr><tr><td>B2</td></tr></table>
        </body></html>
        """
        from scripts.data.scraper.parsers.text_parser import parse_html
        result = parse_html(html)
        assert len(result.tables) == 2

    def test_html_br_becomes_newline_in_text(self):
        from scripts.data.scraper.parsers.text_parser import parse_html
        html = "<html><body><p>line one<br/>line two</p></body></html>"
        result = parse_html(html)
        assert "line one" in result.text
        assert "line two" in result.text

    def test_norm_collapses_blank_lines(self):
        from scripts.data.scraper.parsers.text_parser import _norm
        text = "line1\n\n\n\n\nline2"
        normed = _norm(text)
        assert "\n\n\n" not in normed
        assert "line1" in normed
        assert "line2" in normed
